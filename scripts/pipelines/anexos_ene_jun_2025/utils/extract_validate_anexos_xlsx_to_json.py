#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
extract_validate_anexos_xlsx_to_json.py

Objetivo
--------
Leer anexos_input.xlsx, extraer la información relevante del período Enero-Junio 2025
desde una estructura tabular explícita, y construir un JSON estructurado equivalente
al que antes se generaba desde ANEXOS.pdf.

Salida principal
----------------
Siempre escribe un JSON completo en:

    inputs/raw/anexos_ene_jun_2025/anexos_ene_jun_2025_extraido_desde_xlsx.json

Estructura esperada del XLSX
----------------------------
Workbook: anexos_input.xlsx

Sheets esperadas:

1) lineas
   columnas:
       linea | facturacion_pct | margen_pct

2) zonas
   columnas:
       zona | ventas_usd | ventas_pct

3) totales
   columnas:
       facturacion_usd | margen_usd

4) comentarios
   columnas:
       comentario

Ubicación esperada
------------------
inputs/static_company/documentos_empresa/anexos_input.xlsx

Dependencias
------------
    pip install openpyxl

Uso
---
python scripts/pipelines/anexos_ene_jun_2025/utils/extract_validate_anexos_xlsx_to_json.py
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook


# =============================================================================
# UTILIDADES DE PROYECTO
# =============================================================================
def find_project_root(start_path: Path) -> Path:
    """
    Busca hacia arriba una raíz de proyecto razonable.
    Criterio: presencia simultánea de README.md y requirements.txt.
    """
    candidate = start_path.parent
    for _ in range(10):
        if (candidate / "README.md").exists() and (candidate / "requirements.txt").exists():
            return candidate
        if candidate.parent == candidate:
            break
        candidate = candidate.parent
    raise RuntimeError(
        f"No se pudo localizar la raíz del proyecto desde: {start_path}"
    )


PROJECT_ROOT = find_project_root(Path(__file__).resolve())

INPUTS_RAW = PROJECT_ROOT / "inputs" / "raw" / "anexos_ene_jun_2025"
# OUTPUTS_AUDIT = PROJECT_ROOT / "outputs" / "audit"

EXTRACTED_JSON = INPUTS_RAW / "anexos_ene_jun_2025_extraido_desde_xlsx.json"

XLSX_CANDIDATES = [
    PROJECT_ROOT / "inputs" / "static_company" / "documentos_empresa" / "anexos_input.xlsx",
    PROJECT_ROOT / "inputs" / "static_company" / "documentos_empresa" / "Anexos_input.xlsx",
    PROJECT_ROOT / "inputs" / "static_company" / "documentos_empresa" / "ANEXOS_INPUT.xlsx",
]


# =============================================================================
# UTILIDADES GENERALES
# =============================================================================
def ensure_dirs() -> None:
    INPUTS_RAW.mkdir(parents=True, exist_ok=True)
    # OUTPUTS_AUDIT.mkdir(parents=True, exist_ok=True)


def pretty_json(obj: Any) -> str:
    return json.dumps(obj, ensure_ascii=False, indent=2)


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def coerce_float(value: Any, field_name: str) -> float:
    if value is None or str(value).strip() == "":
        raise ValueError(f"Valor vacío en campo requerido: {field_name}")

    if isinstance(value, (int, float)):
        return float(value)

    s = str(value).strip().replace("%", "").replace(",", ".")
    return float(s)


def coerce_int(value: Any, field_name: str) -> int:
    if value is None or str(value).strip() == "":
        raise ValueError(f"Valor vacío en campo requerido: {field_name}")

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        return int(round(value))

    s = str(value).strip()
    s = s.replace("US$", "").replace("$", "").strip()
    s = s.replace(".", "").replace(",", "")
    return int(s)


def coerce_str(value: Any, field_name: str) -> str:
    if value is None:
        raise ValueError(f"Valor vacío en campo requerido: {field_name}")
    s = str(value).strip()
    if not s:
        raise ValueError(f"Valor vacío en campo requerido: {field_name}")
    return s


def find_xlsx() -> Path:
    for path in XLSX_CANDIDATES:
        if path.exists():
            return path
    raise FileNotFoundError(
        "No se encontró anexos_input.xlsx en rutas esperadas.\n"
        "Colócalo en una de estas ubicaciones:\n- "
        + "\n- ".join(str(p) for p in XLSX_CANDIDATES)
    )


def workbook_headers(ws) -> List[str]:
    """
    Lee la primera fila como headers.
    """
    return [normalize_header(cell.value) for cell in ws[1]]


def require_sheet(wb, sheet_name: str):
    if sheet_name not in wb.sheetnames:
        raise KeyError(
            f"No existe la hoja requerida '{sheet_name}'. "
            f"Hojas disponibles: {wb.sheetnames}"
        )
    return wb[sheet_name]


def validate_headers(actual: List[str], expected: List[str], sheet_name: str) -> None:
    if actual[:len(expected)] != expected:
        raise ValueError(
            f"Headers inválidos en hoja '{sheet_name}'.\n"
            f"Esperado: {expected}\n"
            f"Encontrado: {actual}"
        )


# =============================================================================
# EXTRACCIÓN DESDE XLSX
# =============================================================================
def extract_lineas(ws) -> List[Dict[str, Any]]:
    expected_headers = ["linea", "facturacion_pct", "margen_pct"]
    actual_headers = workbook_headers(ws)
    validate_headers(actual_headers, expected_headers, "lineas")

    rows: List[Dict[str, Any]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        if all(cell is None or str(cell).strip() == "" for cell in row[:3]):
            continue

        linea = coerce_str(row[0], "linea")
        facturacion_pct = coerce_float(row[1], "facturacion_pct")
        margen_pct = coerce_float(row[2], "margen_pct")

        rows.append({
            "linea": linea,
            "facturacion_pct": facturacion_pct,
            "margen_pct": margen_pct,
        })

    if not rows:
        raise ValueError("La hoja 'lineas' no contiene filas de datos válidas.")

    return rows


def extract_zonas(ws) -> List[Dict[str, Any]]:
    expected_headers = ["zona", "ventas_usd", "ventas_pct"]
    actual_headers = workbook_headers(ws)
    validate_headers(actual_headers, expected_headers, "zonas")

    rows: List[Dict[str, Any]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        if all(cell is None or str(cell).strip() == "" for cell in row[:3]):
            continue

        zona = coerce_str(row[0], "zona")
        ventas_usd = coerce_int(row[1], "ventas_usd")
        ventas_pct = coerce_float(row[2], "ventas_pct")

        rows.append({
            "zona": zona,
            "ventas_usd": ventas_usd,
            "ventas_pct": ventas_pct,
        })

    if not rows:
        raise ValueError("La hoja 'zonas' no contiene filas de datos válidas.")

    return rows


def extract_totales(ws) -> Dict[str, int]:
    expected_headers = ["facturacion_usd", "margen_usd"]
    actual_headers = workbook_headers(ws)
    validate_headers(actual_headers, expected_headers, "totales")

    first_data_row = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        if all(cell is None or str(cell).strip() == "" for cell in row[:2]):
            continue
        first_data_row = row
        break

    if first_data_row is None:
        raise ValueError("La hoja 'totales' no contiene una fila válida de datos.")

    return {
        "facturacion_usd": coerce_int(first_data_row[0], "facturacion_usd"),
        "margen_usd": coerce_int(first_data_row[1], "margen_usd"),
    }


def extract_comments(ws) -> List[str]:
    expected_headers = ["comentario"]
    actual_headers = workbook_headers(ws)
    validate_headers(actual_headers, expected_headers, "comentarios")

    comments: List[str] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        if len(row) == 0:
            continue
        value = row[0]
        if value is None or str(value).strip() == "":
            continue
        comments.append(str(value).strip())

    return comments


def build_payload_from_xlsx(xlsx_path: Path) -> Dict[str, Any]:
    wb = load_workbook(xlsx_path, data_only=True)

    ws_lineas = require_sheet(wb, "lineas")
    ws_zonas = require_sheet(wb, "zonas")
    ws_totales = require_sheet(wb, "totales")
    ws_comentarios = require_sheet(wb, "comentarios")

    lineas = extract_lineas(ws_lineas)
    zonas = extract_zonas(ws_zonas)
    totales = extract_totales(ws_totales)
    comentarios = extract_comments(ws_comentarios)

    payload = {
        "metadata": {
            "fuente": xlsx_path.name,
            "periodo": "Enero-Junio 2025",
            "generado_desde_xlsx": True,
        },
        "totales": {
            "facturacion_usd": totales["facturacion_usd"],
            "margen_usd": totales["margen_usd"],
        },
        "lineas": lineas,
        "zonas": zonas,
        "comentarios_extraidos_pdf": comentarios,
    }

    return payload


# =============================================================================
# VALIDACIONES DE CONSISTENCIA BÁSICA
# =============================================================================
def validate_payload(payload: Dict[str, Any]) -> None:
    required_top = ["metadata", "totales", "lineas", "zonas", "comentarios_extraidos_pdf"]
    for key in required_top:
        if key not in payload:
            raise KeyError(f"Falta la clave requerida en el payload: '{key}'")

    if not isinstance(payload["lineas"], list) or len(payload["lineas"]) == 0:
        raise ValueError("'lineas' debe ser una lista no vacía.")

    if not isinstance(payload["zonas"], list) or len(payload["zonas"]) == 0:
        raise ValueError("'zonas' debe ser una lista no vacía.")

    if not isinstance(payload["comentarios_extraidos_pdf"], list):
        raise ValueError("'comentarios_extraidos_pdf' debe ser una lista.")

    for row in payload["lineas"]:
        for key in ["linea", "facturacion_pct", "margen_pct"]:
            if key not in row:
                raise KeyError(f"Falta la clave '{key}' en una fila de 'lineas'.")

    for row in payload["zonas"]:
        for key in ["zona", "ventas_usd", "ventas_pct"]:
            if key not in row:
                raise KeyError(f"Falta la clave '{key}' en una fila de 'zonas'.")

    for key in ["facturacion_usd", "margen_usd"]:
        if key not in payload["totales"]:
            raise KeyError(f"Falta la clave '{key}' en 'totales'.")


# =============================================================================
# MAIN
# =============================================================================
def main() -> None:
    ensure_dirs()

    print("=" * 100)
    print("EXTRACCIÓN DE ANEXOS XLSX -> JSON")
    print("=" * 100)
    print(f"Project root      : {PROJECT_ROOT.resolve()}")

    xlsx_path = find_xlsx()
    print(f"XLSX fuente       : {xlsx_path.resolve()}")
    print(f"JSON de salida    : {EXTRACTED_JSON.resolve()}")

    payload = build_payload_from_xlsx(xlsx_path)
    validate_payload(payload)

    print("\n" + "=" * 100)
    print("PAYLOAD EXTRAÍDO DESDE XLSX")
    print("=" * 100)
    print(pretty_json(payload))

    with open(EXTRACTED_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f"\n[OK] JSON extraído guardado en: {EXTRACTED_JSON.resolve()}")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        sys.exit(1)
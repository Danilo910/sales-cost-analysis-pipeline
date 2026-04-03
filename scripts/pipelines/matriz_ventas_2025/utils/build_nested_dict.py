#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from pathlib import Path

# -------------------------------------------------
# PROJECT PATHS
# -------------------------------------------------
def find_project_root(start_path: Path) -> Path:
    """
    Busca hacia arriba una raíz de proyecto razonable.
    Criterio: presencia simultánea de README.md y requirements.txt.
    """
    candidate = start_path.parent
    for _ in range(10):
        if (candidate / "README.md").exists() and (candidate / "requirements.txt").exists():
            return candidate
        if candidate.parent == candidate:
            break
        candidate = candidate.parent
    raise RuntimeError(
        f"No se pudo localizar la raíz del proyecto desde: {start_path}"
    )

PROJECT_ROOT = find_project_root(Path(__file__).resolve())
INPUTS_STATIC_COMPANY = PROJECT_ROOT / "inputs" / "static_company" / "documentos_empresa"
INPUTS_RAW_SHARED = PROJECT_ROOT / "inputs" / "raw" / "shared"
import re
import json
import openpyxl

FILE_NAME = "Costos_Vs_ventas_hasta_Diciembre_2025.xlsx"
SHEET_NAME = None  # None = primera hoja

MONTHS = [
    "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SETIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"
]
EXTRA_TOTALS = ["TOTAL COSTO", "TOTAL VENTA"]

# Si en tu Excel usan "SETIEMBRE" o "SEPTIEMBRE", acepta ambos
MONTH_ALIASES = {
    "SEPTIEMBRE": "SETIEMBRE"
}


def norm_str(x):
    if x is None:
        return ""
    if isinstance(x, str):
        return x.strip()
    return str(x).strip()


def safe_num(x):
    if x is None or x == "":
        return 0
    try:
        return float(x)
    except Exception:
        s = str(x).replace(",", "").strip()
        try:
            return float(s)
        except Exception:
            return 0


def normalize_month_label(label: str) -> str:
    label = label.upper().strip()
    return MONTH_ALIASES.get(label, label)


def find_blocks(ws):
    """
    Encuentra bloques de productos. Soporta títulos del tipo:
        - 'PRODUCTOS DE <LINEA>'
        - 'OTROS PRODUCTOS'
        - 'OTRAS EMPRESAS'
        - 'PRODUCTOS DE OTRAS EMPRESAS'

    Retorna lista de dicts: [{"linea": ..., "row": r, "col": c, "raw": ...}, ...]
    """
    blocks = []
    max_r = ws.max_row
    max_c = ws.max_column

    pat_productos_de = re.compile(r"^\s*PRODUCTOS\s+DE\s+(.+?)\s*$", re.IGNORECASE)
    pat_otros_productos = re.compile(r"^\s*OTROS\s+PRODUCTOS\s*$", re.IGNORECASE)
    pat_otras_empresas  = re.compile(r"^\s*OTRAS?\s+EMPRESAS?\s*$", re.IGNORECASE)
    pat_prod_otras_emp  = re.compile(r"^\s*PRODUCTOS\s+DE\s+OTRAS?\s+EMPRESAS?\s*$", re.IGNORECASE)

    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue

            s = v.strip()
            if not s:
                continue

            m = pat_productos_de.match(s)
            if m:
                linea = m.group(1).strip()
                # normaliza algunos nombres comunes
                # if linea in ["OTRAS EMPRESAS", "OTRA EMPRESA"]:
                #     linea = "OTRAS EMPRESAS"
                blocks.append({"linea": linea, "row": r, "col": c, "raw": s})
                continue

            if pat_prod_otras_emp.match(s) or pat_otras_empresas.match(s):
                blocks.append({"linea": "OTRAS EMPRESAS", "row": r, "col": c, "raw": s})
                continue

            if pat_otros_productos.match(s):
                blocks.append({"linea": "OTROS PRODUCTOS", "row": r, "col": c, "raw": s})
                continue

    # Elimina duplicados (por si hay celdas combinadas o repetidas)
    seen = set()
    uniq = []
    for b in blocks:
        key = (b["linea"], b["row"], b["col"])
        if key not in seen:
            uniq.append(b)
            seen.add(key)
    return uniq


def find_header_row(ws, start_row, max_scan=15):
    """
    Desde start_row (encabezado del bloque), busca la fila donde aparezcan
    'DESCRIPCION' y 'PRECIO COSTO'.
    """
    for r in range(start_row, start_row + max_scan + 1):
        row_vals = [norm_str(ws.cell(r, c).value).upper() for c in range(1, ws.max_column + 1)]
        if "DESCRIPCION" in row_vals and "PRECIO COSTO" in row_vals:
            return r
    return None


def build_colmap(ws, header_row):
    top = [norm_str(ws.cell(header_row, c).value).upper() for c in range(1, ws.max_column + 1)]
    sub = [norm_str(ws.cell(header_row + 1, c).value).upper() for c in range(1, ws.max_column + 1)]

    descripcion_col = None
    precio_costo_col = None
    for idx, v in enumerate(top, start=1):
        if v == "DESCRIPCION":
            descripcion_col = idx
        if v == "PRECIO COSTO":
            precio_costo_col = idx

    if descripcion_col is None or precio_costo_col is None:
        raise ValueError("No se encontró DESCRIPCION o PRECIO COSTO en la cabecera.")

    month_to_cols = {}
    totals_to_cols = {}

    for c in range(1, ws.max_column + 1):
        label = normalize_month_label(top[c - 1])
        if label in MONTHS:
            month_to_cols[label] = (c, c + 1)

    for c in range(1, ws.max_column + 1):
        label = top[c - 1]
        if label in EXTRA_TOTALS:
            key = label.replace(" ", "_")  # TOTAL_COSTO / TOTAL_VENTA
            totals_to_cols[key] = (c, c + 1)

    return {
        "descripcion_col": descripcion_col,
        "precio_costo_col": precio_costo_col,
        "month_to_cols": month_to_cols,
        "totals_to_cols": totals_to_cols,
    }


def parse_products_in_block(ws, block, colmap):
    """
    Lee filas de productos desde header_row+2 hacia abajo hasta que:
    - encuentre otro encabezado de bloque (PRODUCTOS DE..., OTROS PRODUCTOS, OTRAS EMPRESAS)
    - o encuentre 'VALORES EXPRESADOS...' (pie)
    - o haya filas vacías consecutivas
    """
    data = {}
    r = block["header_row"] + 2
    max_r = ws.max_row

    # Detectores de próximo bloque
    pat_next = re.compile(
        r"^\s*(PRODUCTOS\s+DE\s+.+|OTROS\s+PRODUCTOS|OTRAS?\s+EMPRESAS|PRODUCTOS\s+DE\s+OTRAS?\s+EMPRESAS)\s*$",
        re.IGNORECASE
    )

    empty_streak = 0

    while r <= max_r:
        # 1) ¿hay encabezado de próximo bloque en la fila?
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and pat_next.match(v.strip()):
                return data

        desc = ws.cell(r, colmap["descripcion_col"]).value
        desc_s = norm_str(desc)

        if desc_s == "":
            empty_streak += 1
            if empty_streak >= 2:
                return data
            r += 1
            continue
        empty_streak = 0

        if "VALORES EXPRESADOS" in desc_s.upper():
            return data

        producto = desc_s
        precio = safe_num(ws.cell(r, colmap["precio_costo_col"]).value)

        prod_dict = {"precio_costo": precio}

        for mes, (cu, cv) in colmap["month_to_cols"].items():
            unidades = safe_num(ws.cell(r, cu).value)
            valor = safe_num(ws.cell(r, cv).value)
            # unidades puede ser float por Excel; lo convertimos a int si es entero
            if float(unidades).is_integer():
                unidades = int(unidades)
            prod_dict[mes] = {"unidades": unidades, "valor": valor}

        for tot_key, (cu, cv) in colmap["totals_to_cols"].items():
            unidades = safe_num(ws.cell(r, cu).value)
            valor = safe_num(ws.cell(r, cv).value)
            if float(unidades).is_integer():
                unidades = int(unidades)
            prod_dict[tot_key] = {"unidades": unidades, "valor": valor}

        data[producto] = prod_dict
        r += 1

    return data


def main():
    base = PROJECT_ROOT
    path = INPUTS_STATIC_COMPANY / FILE_NAME
    INPUTS_RAW_SHARED.mkdir(parents=True, exist_ok=True)
    print("=== Verificación de ruta ===")
    print("Directorio actual:", base)
    print("Archivo esperado:", path)
    if not path.exists():
        raise FileNotFoundError(f"No encuentro el archivo: {path}")

    print("\n=== Cargando Excel ===")
    wb = openpyxl.load_workbook(path, data_only=True)
    print("Hojas encontradas:", wb.sheetnames)

    ws = wb[SHEET_NAME] if SHEET_NAME else wb[wb.sheetnames[0]]
    print("Usando hoja:", ws.title)

    print("\n=== Buscando bloques de líneas de producto ===")
    blocks = find_blocks(ws)
    print(f"Bloques encontrados: {len(blocks)}")
    for b in blocks:
        print(f" - {b['raw']}  -> linea='{b['linea']}' (fila {b['row']}, col {b['col']})")

    if not blocks:
        raise RuntimeError("No encontré ningún bloque. Revisa el formato del Excel.")

    data = {}
    for b in blocks:
        header_row = find_header_row(ws, b["row"], max_scan=15)
        if header_row is None:
            print(f"\n[ADVERTENCIA] No encontré cabecera DESCRIPCION/PRECIO COSTO para '{b['linea']}' (desde fila {b['row']}).")
            continue

        b["header_row"] = header_row
        colmap = build_colmap(ws, header_row)

        print(f"\n=== Procesando bloque: {b['linea']} ===")
        print(f"Cabecera en fila: {header_row}")
        print("Meses detectados:", list(colmap["month_to_cols"].keys()))
        print("Totales detectados:", list(colmap["totals_to_cols"].keys()))

        productos = parse_products_in_block(ws, b, colmap)
        data[b["linea"]] = productos

        print(f"Productos leídos en {b['linea']}: {len(productos)}")
        if productos:
            first_prod = next(iter(productos))
            sample = productos[first_prod]
            print("Ejemplo producto:", first_prod)
            print("  precio_costo:", sample.get("precio_costo"))
            for mes in ["ENERO", "FEBRERO", "MARZO"]:
                if mes in sample:
                    print(f"  {mes}:", sample[mes])

    print("\n=== Resumen final ===")
    for linea, prods in data.items():
        print(f"- {linea}: {len(prods)} productos")

    out_json = INPUTS_RAW_SHARED / "diccionario_ventas_costos_2025.json"
    with open(out_json, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print("\nJSON guardado en:", out_json)


if __name__ == "__main__":
    main()
